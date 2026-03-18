import openpyxl
import unicodedata
import re
from pathlib import Path


def fix_encoding(s: str) -> str:
    """repair Mac Roman mojibake (UTF-8 bytes misread as Mac Roman)."""
    try:
        return s.encode("mac_roman").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        return s


def normalize(s: str) -> str:
    """lowercase, strip accents, collapse whitespace."""
    s = s.lower().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s


def main(
    mathnli_path: str = "mathnli.xlsx",
    extracted_path: str = "extracted.xlsx",
    output_path: str = "mathnli_merged.xlsx",
    report_path: str = "unmatched_report.xlsx",
):
    # load mathnli
    wb1 = openpyxl.load_workbook(mathnli_path)
    ws1 = wb1.active
    headers = [cell.value for cell in ws1[1]]
    num_db_cols = len(headers) - 1  # everything after the term column

    mathnli_data = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        mathnli_data.append(list(row))

    # load extracted
    wb2 = openpyxl.load_workbook(extracted_path)
    ws2 = wb2.active
    extracted_raw = [
        str(row[0])
        for row in ws2.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    ]

    # fix encoding, build lookup
    extracted_fixed = [fix_encoding(t) for t in extracted_raw]

    mathnli_norm_map = {}
    for i, row in enumerate(mathnli_data):
        mathnli_norm_map[normalize(str(row[0]))] = i

    # match
    matched_indices = set()
    unmatched = []

    for raw, fixed in zip(extracted_raw, extracted_fixed):
        norm = normalize(fixed)
        if norm in mathnli_norm_map:
            matched_indices.add(mathnli_norm_map[norm])
        else:
            unmatched.append((fixed, raw != fixed))

    # write merged output
    out_headers = headers + ["extracted", "match_status"]
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(out_headers)

    for i, row in enumerate(mathnli_data):
        ext_val = 1 if i in matched_indices else 0
        status = "matched" if i in matched_indices else ""
        ws_out.append(row + [ext_val, status])

    for term, was_repaired in unmatched:
        status = "unmatched_new_encoding_repaired" if was_repaired else "unmatched_new"
        ws_out.append([term] + [0] * num_db_cols + [1, status])

    wb_out.save(output_path)

    # write unmatched report
    wb_report = openpyxl.Workbook()
    ws_report = wb_report.active
    ws_report.append(["unmatched_term", "encoding_repaired"])
    for term, was_repaired in unmatched:
        ws_report.append([term, "yes" if was_repaired else "no"])
    wb_report.save(report_path)

    # summary stuff
    repaired_count = sum(1 for _, r in unmatched if r)
    print(f"Original rows:         {len(mathnli_data)}")
    print(f"Extracted terms:       {len(extracted_fixed)}")
    print(f"Matched:               {len(matched_indices)}")
    print(f"Unmatched (appended):  {len(unmatched)}")
    print(f"  encoding-repaired:   {repaired_count}")
    print(f"Final rows:            {len(mathnli_data) + len(unmatched)}")
    print(f"\nWrote: {output_path}")
    print(f"Wrote: {report_path}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mathnli", default="mathnli.xlsx", help="Path to main term matrix")
    parser.add_argument("--extracted", default="extracted.xlsx", help="Path to extracted term list")
    parser.add_argument("--output", default="mathnli_merged.xlsx", help="Output merged file")
    parser.add_argument("--report", default="unmatched_report.xlsx", help="Output unmatched report")
    args = parser.parse_args()

    main(args.mathnli, args.extracted, args.output, args.report)
